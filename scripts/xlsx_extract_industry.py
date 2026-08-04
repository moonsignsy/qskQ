import zipfile, re, datetime
from xml.etree import ElementTree as ET
from collections import defaultdict

M = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
NS = {'a': M}

def col_to_idx(col):
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1

def load(path):
    z = zipfile.ZipFile(path)
    shared = []
    if 'xl/sharedStrings.xml' in z.namelist():
        root = ET.fromstring(z.read('xl/sharedStrings.xml'))
        for si in root.findall('a:si', NS):
            shared.append(''.join(t.text or '' for t in si.iter('{'+M+'}t')))
    wb = ET.fromstring(z.read('xl/workbook.xml'))
    rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    relmap = {r.get('Id'): r.get('Target') for r in rels}
    sheets = []
    for s in wb.find('a:sheets', NS).findall('a:sheet', NS):
        rid = s.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        target = relmap[rid]
        if not target.startswith('xl/'):
            target = 'xl/' + target
        sheets.append((s.get('name'), target))
    return z, shared, sheets

def read_full(z, shared, target):
    root = ET.fromstring(z.read(target))
    data = root.find('a:sheetData', NS)
    out = []
    for row in data.findall('a:row', NS):
        cells = {}
        maxc = -1
        for c in row.findall('a:c', NS):
            col = re.match(r'[A-Z]+', c.get('r')).group()
            ci = col_to_idx(col)
            t = c.get('t')
            val = None
            if t == 's':
                v = c.find('a:v', NS)
                if v is not None:
                    val = shared[int(v.text)]
            elif t == 'inlineStr':
                is_ = c.find('a:is', NS)
                if is_ is not None:
                    val = ''.join(tt.text or '' for tt in is_.iter('{'+M+'}t'))
            else:
                v = c.find('a:v', NS)
                if v is not None:
                    val = v.text  # keep raw (number or string)
            cells[ci] = val
            maxc = max(maxc, ci)
        out.append([cells.get(i, None) for i in range(maxc + 1)])
    return out

def norm(x):
    return '' if x is None else str(x).strip()

def to_num(x):
    if x is None:
        return None
    s = str(x).replace(',', '').strip()
    try:
        return float(s)
    except Exception:
        return None

def serial_to_date(x):
    n = to_num(x)
    if n is None:
        return norm(x)
    if 40000 <= n <= 70000:
        try:
            d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(round(n)))
            return d.isoformat()
        except Exception:
            return norm(x)
    return norm(x)

# ---- per-sheet column maps (0-based) ----
# fields: 序号,客户姓名,手机号码,区域,行业,企业形式,来源渠道,跟进结果,意向,签单项目,
#         签单金额,收款金额,成本,业绩金额,成单周期,签单日期,首次联系日期
# value None = not present in that sheet
SHEET_MAP = {
    '嘉乐': dict(hdr=2, date={1,16},
        cols=dict(序号=0,客户姓名=2,手机号码=3,区域=4,行业=5,企业形式=6,来源渠道=7,跟进结果=8,
                  意向=9,签单项目=21,签单金额=17,收款金额=18,成本=19,业绩金额=20,成单周期=22,签单日期=16,首次联系日期=1)),
    '谨谨': dict(hdr=2, date={1,19},
        cols=dict(序号=0,客户姓名=2,手机号码=3,区域=4,行业=5,企业形式=6,来源渠道=7,跟进结果=8,
                  意向=9,签单项目=None,签单金额=20,收款金额=21,成本=22,业绩金额=None,成单周期=None,签单日期=19,首次联系日期=1)),
    '2026': dict(hdr=2, date={2,15},
        cols=dict(序号=0,客户姓名=3,手机号码=4,区域=5,行业=6,企业形式=7,来源渠道=8,跟进结果=9,
                  意向=10,签单项目=20,签单金额=16,收款金额=17,成本=18,业绩金额=19,成单周期=21,签单日期=15,首次联系日期=2)),
    '交换资源': dict(hdr=2, date={2,15},
        cols=dict(序号=0,客户姓名=3,手机号码=4,区域=5,行业=6,企业形式=7,来源渠道=8,跟进结果=9,
                  意向=10,签单项目=20,签单金额=16,收款金额=17,成本=18,业绩金额=19,成单周期=21,签单日期=15,首次联系日期=2)),
    '已退款': dict(hdr=1, date={1,9},
        cols=dict(序号=0,客户姓名=2,手机号码=3,区域=4,行业=5,企业形式=6,来源渠道=7,跟进结果=8,
                  意向=None,签单项目=14,签单金额=None,收款金额=11,成本=12,业绩金额=13,成单周期=15,签单日期=9,首次联系日期=1)),
    '已成单': dict(hdr=1, date={2},
        cols=dict(序号=0,客户姓名=4,手机号码=5,区域=6,行业=7,企业形式=None,来源渠道=8,跟进结果=9,
                  意向=None,签单项目=14,签单金额=10,收款金额=11,成本=12,业绩金额=13,成单周期=15,签单日期=2,首次联系日期=None)),
}

OUT_FIELDS = ['序号','客户姓名','手机号码','区域','行业','企业形式','来源渠道','跟进结果','意向',
              '签单项目','签单金额','收款金额','成本','业绩金额','成单周期','首次联系日期','签单日期']

SRC = r"C:\Users\EDY\Desktop\苏外组 - 线上资源跟进表.xlsx"
OUT = r"C:\Users\EDY\Desktop\苏外组 - 线上资源跟进表_行业数据.xlsx"

z, shared, sheets = load(SRC)
detail = []
for name, target in sheets:
    if name not in SHEET_MAP:
        print(f"[skip] 未配置列映射: {name}")
        continue
    cfg = SHEET_MAP[name]
    rows = read_full(z, shared, target)
    cols = cfg['cols']
    datecols = cfg['date']
    start = cfg['hdr'] + 1
    for row in rows[start:]:
        has = any(norm(row[c]) != '' for c in (cols.get('序号'), cols.get('客户姓名'), cols.get('行业'), cols.get('手机号码')) if c is not None and c < len(row))
        if not has:
            continue
        rec = {'负责人/工作表': name}
        # detect which original column index each field maps to
        for f in OUT_FIELDS:
            ci = cols.get(f)
            if ci is None or ci >= len(row):
                rec[f] = ''
            else:
                v = row[ci]
                if f in ('签单金额','收款金额','成本','业绩金额'):
                    rec[f] = to_num(v)
                elif ci in datecols:
                    rec[f] = serial_to_date(v)
                else:
                    rec[f] = norm(v)
        detail.append(rec)

print("提取明细记录数:", len(detail))

# ---- write output ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws1 = wb.active
ws1.title = '行业明细'
headers = ['负责人/工作表'] + OUT_FIELDS
ws1.append(headers)
for rec in detail:
    ws1.append([rec[h] for h in headers])

hdr_fill = PatternFill('solid', fgColor='2563EB')
for cell in ws1[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.freeze_panes = 'A2'
ws1.auto_filter.ref = f"A1:{chr(64+len(headers))}{len(detail)+1}"

# ---- summary ----
ws2 = wb.create_sheet('行业汇总')
agg = defaultdict(lambda: {'记录数':0,'签单金额':0.0,'收款金额':0.0,'业绩金额':0.0})
for rec in detail:
    ind = rec['行业'] if rec['行业'] else '(未填写)'
    a = agg[ind]
    a['记录数'] += 1
    for k in ('签单金额','收款金额','业绩金额'):
        v = rec.get(k)
        if isinstance(v,(int,float)):
            a[k] += v
summary_rows = sorted(agg.items(), key=lambda kv: kv[1]['记录数'], reverse=True)
ws2.append(['行业','记录数','签单金额合计','收款金额合计','业绩金额合计'])
for ind, a in summary_rows:
    ws2.append([ind, a['记录数'], round(a['签单金额'],2), round(a['收款金额'],2), round(a['业绩金额'],2)])
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
ws2.freeze_panes = 'A2'

for ws in (ws1, ws2):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(width+2, 10), 42)

wb.save(OUT)
print("已导出:", OUT)
print("明细行数:", len(detail), "| 行业种类(含未填写):", len(agg))
